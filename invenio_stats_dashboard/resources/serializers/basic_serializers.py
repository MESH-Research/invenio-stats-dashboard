# Part of the Invenio-Stats-Dashboard extension for InvenioRDM
# Copyright (C) 2025 Mesh Research
#
# Invenio-Stats-Dashboard is free software; you can redistribute it and/or modify
# it under the terms of the MIT License; see LICENSE file for more details.

"""Serializers for invenio-stats-dashboard content negotiation."""

import csv
import io
import json
import xml.etree.ElementTree as ET
from typing import Any

from flask import Response, current_app
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook


class StatsJSONSerializer:
    """JSON serializer for stats responses."""

    def serialize(self, data: dict | list | bytes, **kwargs) -> dict | list | Any:
        """Serialize data to JSON format.

        Args:
            data: The data to serialize (dict, list, or bytes)
            **kwargs: Additional keyword arguments

        Returns:
            Raw JSON-ready data (dict or list) ready to be dumped to JSON
        """
        if isinstance(data, bytes):
            try:
                json_data = json.loads(data.decode('utf-8'))
                return json_data
            except (json.JSONDecodeError, UnicodeDecodeError) as e:
                current_app.logger.warning(f"Failed to decode bytes data to JSON: {e}")
                return {"error": "Invalid JSON data"}

        return data


class StatsCSVSerializer:
    """CSV serializer for stats responses."""

    def serialize(self, data: dict | list, **kwargs) -> Response:
        """Serialize data to CSV format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with CSV content
        """
        output = io.StringIO()

        if isinstance(data, list) and data:
            # Handle list of records (e.g., daily stats)
            if isinstance(data[0], dict):
                fieldnames = self._get_fieldnames(data[0])
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
        elif isinstance(data, dict):
            # Handle dict responses (e.g., comprehensive stats)
            csv_writer = csv.writer(output)
            for key, value in data.items():
                if isinstance(value, list):
                    csv_writer.writerow([key])
                    if value and isinstance(value[0], dict):
                        # List of dicts - write as sub-table
                        sub_fieldnames = self._get_fieldnames(value[0])
                        sub_writer = csv.DictWriter(output, fieldnames=sub_fieldnames)
                        sub_writer.writeheader()
                        sub_writer.writerows(value)
                    else:
                        for item in value:
                            csv_writer.writerow([item])
                else:
                    csv_writer.writerow([key, value])

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Type": "text/csv; charset=utf-8",
                "Content-Disposition": "attachment; filename=stats.csv",
            },
        )

    def _get_fieldnames(self, record: dict[str, Any]) -> list[str]:
        """Extract fieldnames from a record, handling nested structures.

        Args:
            record: Dictionary record to extract fieldnames from

        Returns:
            List of field names
        """
        fieldnames = []
        for key, value in record.items():
            if isinstance(value, dict):
                # For nested dicts, flatten with dot notation
                for subkey in value.keys():
                    fieldnames.append(f"{key}.{subkey}")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                fieldnames.append(key)
            else:
                fieldnames.append(key)
        return fieldnames


class StatsXMLSerializer:
    """XML serializer for stats responses."""

    def serialize(self, data: dict | list, **kwargs) -> Response:
        """Serialize data to XML format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with XML content
        """
        root = ET.Element("stats")

        if isinstance(data, list):
            for i, item in enumerate(data):
                item_elem = ET.SubElement(root, "item", id=str(i))
                self._dict_to_xml(item, item_elem)
        elif isinstance(data, dict):
            self._dict_to_xml(data, root)

        return Response(
            ET.tostring(root, encoding="unicode"),
            mimetype="application/xml",
            headers={"Content-Type": "application/xml; charset=utf-8"},
        )

    def _dict_to_xml(self, data: dict[str, Any], parent: ET.Element) -> None:
        """Convert dictionary to XML elements.

        Modifies the input xml parent element in place.

        Args:
            data: Dictionary to convert
            parent: Parent XML element
        """
        for key, value in data.items():
            if isinstance(value, dict):
                elem = ET.SubElement(parent, key)
                self._dict_to_xml(value, elem)
            elif isinstance(value, list):
                list_elem = ET.SubElement(parent, key)
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        item_elem = ET.SubElement(list_elem, "item", id=str(i))
                        self._dict_to_xml(item, item_elem)
                    else:
                        item_elem = ET.SubElement(list_elem, "item", id=str(i))
                        item_elem.text = str(item)
            else:
                elem = ET.SubElement(parent, key)
                elem.text = str(value)


class StatsExcelSerializer:
    """Excel serializer for stats responses."""

    def serialize(self, data: dict | list, **kwargs) -> Response:
        """Serialize data to Excel format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with Excel content
        """
        wb = Workbook()

        if isinstance(data, list) and data:
            # Handle list of records
            ws = wb.active
            if ws is not None and hasattr(ws, "title"):
                ws.title = "Statistics Data"
            self._list_to_excel(data, ws)
        elif isinstance(data, dict):
            # Handle dict response - create multiple sheets
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    # List of records - create sheet
                    ws = wb.create_sheet(title=str(key.replace("_", " ").title()))
                    self._list_to_excel(value, ws)
                elif isinstance(value, dict):
                    # Nested dict - create sheet
                    ws = wb.create_sheet(title=str(key.replace("_", " ").title()))
                    self._dict_to_excel(value, ws)
                else:
                    # Simple value - add to summary sheet
                    if "Summary" not in [sheet.title for sheet in wb.worksheets]:
                        ws = wb.create_sheet(title="Summary")
                    else:
                        ws = wb["Summary"]
                    ws.append([key, value])

        # Style the worksheets
        for ws in wb.worksheets:
            self._style_worksheet(ws)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Type": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                "Content-Disposition": "attachment; filename=stats.xlsx",
            },
        )

    def _list_to_excel(self, data: list[dict[str, Any]], ws) -> None:
        """Convert list of records to Excel worksheet.

        Args:
            data: List of records
            ws: Excel worksheet
        """
        if not data:
            return

        # Get fieldnames
        fieldnames = self._get_fieldnames(data[0])

        # Write headers
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)

        # Write data
        for row, record in enumerate(data, 2):
            for col, field in enumerate(fieldnames, 1):
                value = self._get_nested_value(record, field)
                ws.cell(row=row, column=col, value=value)

    def _dict_to_excel(self, data: dict[str, Any], ws) -> None:
        """Convert dictionary to Excel worksheet.

        Args:
            data: Dictionary to convert
            ws: Excel worksheet
        """
        for row, (key, value) in enumerate(data.items(), 1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))

    def _get_fieldnames(self, record: dict[str, Any]) -> list[str]:
        """Extract fieldnames from a record, handling nested structures.

        Args:
            record: Dictionary record to extract fieldnames from

        Returns:
            List of field names
        """
        fieldnames = []
        for key, value in record.items():
            if isinstance(value, dict):
                # For nested dicts, flatten with dot notation
                for subkey in value.keys():
                    fieldnames.append(f"{key}.{subkey}")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                # For lists of dicts, use the key as fieldname
                fieldnames.append(key)
            else:
                fieldnames.append(key)
        return fieldnames

    def _get_nested_value(self, record: dict[str, Any], field: str) -> str:
        """Get value from nested field using dot notation.

        Args:
            record: Dictionary record
            field: Field name (may contain dots for nesting)

        Returns:
            String value
        """
        if "." in field:
            keys = field.split(".")
            value = record
            for key in keys:
                if isinstance(value, dict) and key in value:
                    value = value[key]
                else:
                    return ""
            return str(value) if value is not None else ""
        else:
            return str(record.get(field, ""))

    def _style_worksheet(self, ws) -> None:
        """Apply styling to Excel worksheet.

        Args:
            ws: Excel worksheet to style
        """
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
