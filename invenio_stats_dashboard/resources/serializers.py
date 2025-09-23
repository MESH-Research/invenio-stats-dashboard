# Part of the Invenio-Stats-Dashboard extension for InvenioRDM
# Copyright (C) 2025 Mesh Research
#
# Invenio-Stats-Dashboard is free software; you can redistribute it and/or modify
# it under the terms of the MIT License; see LICENSE file for more details.

"""Serializers for invenio-stats-dashboard content negotiation."""

import csv
import io
import xml.etree.ElementTree as ET
from typing import Any, Dict, List, Union

from flask import Response
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook


class StatsJSONSerializer:
    """JSON serializer for stats responses."""

    def serialize(self, data: Union[Dict, List], **kwargs) -> Union[Dict, List]:
        """Return data directly for Flask to handle JSON serialization.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            The data directly - Flask will handle JSON serialization
        """
        return data


class StatsCSVSerializer:
    """CSV serializer for stats responses."""

    def serialize(self, data: Union[Dict, List], **kwargs) -> Response:
        """Serialize data to CSV format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with CSV content
        """
        output = io.StringIO()

        if isinstance(data, list) and data:
            # Handle list of records (e.g., daily stats)
            if isinstance(data[0], dict):
                fieldnames = self._get_fieldnames(data[0])
                writer = csv.DictWriter(output, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(data)
        elif isinstance(data, dict):
            # Handle dict responses (e.g., comprehensive stats)
            csv_writer = csv.writer(output)
            for key, value in data.items():
                if isinstance(value, list):
                    csv_writer.writerow([key])
                    if value and isinstance(value[0], dict):
                        # List of dicts - write as sub-table
                        sub_fieldnames = self._get_fieldnames(value[0])
                        sub_writer = csv.DictWriter(output, fieldnames=sub_fieldnames)
                        sub_writer.writeheader()
                        sub_writer.writerows(value)
                    else:
                        for item in value:
                            csv_writer.writerow([item])
                else:
                    csv_writer.writerow([key, value])

        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Type": "text/csv; charset=utf-8",
                "Content-Disposition": "attachment; filename=stats.csv",
            },
        )

    def _get_fieldnames(self, record: Dict[str, Any]) -> List[str]:
        """Extract fieldnames from a record, handling nested structures.

        Args:
            record: Dictionary record to extract fieldnames from

        Returns:
            List of field names
        """
        fieldnames = []
        for key, value in record.items():
            if isinstance(value, dict):
                # For nested dicts, flatten with dot notation
                for subkey in value.keys():
                    fieldnames.append(f"{key}.{subkey}")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                fieldnames.append(key)
            else:
                fieldnames.append(key)
        return fieldnames


class StatsXMLSerializer:
    """XML serializer for stats responses."""

    def serialize(self, data: Union[Dict, List], **kwargs) -> Response:
        """Serialize data to XML format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with XML content
        """
        root = ET.Element("stats")

        if isinstance(data, list):
            for i, item in enumerate(data):
                item_elem = ET.SubElement(root, "item", id=str(i))
                self._dict_to_xml(item, item_elem)
        elif isinstance(data, dict):
            self._dict_to_xml(data, root)

        return Response(
            ET.tostring(root, encoding="unicode"),
            mimetype="application/xml",
            headers={"Content-Type": "application/xml; charset=utf-8"},
        )

    def _dict_to_xml(self, data: Dict[str, Any], parent: ET.Element) -> None:
        """Convert dictionary to XML elements.

        Modifies the input xml parent element in place.

        Args:
            data: Dictionary to convert
            parent: Parent XML element
        """
        for key, value in data.items():
            if isinstance(value, dict):
                elem = ET.SubElement(parent, key)
                self._dict_to_xml(value, elem)
            elif isinstance(value, list):
                list_elem = ET.SubElement(parent, key)
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        item_elem = ET.SubElement(list_elem, "item", id=str(i))
                        self._dict_to_xml(item, item_elem)
                    else:
                        item_elem = ET.SubElement(list_elem, "item", id=str(i))
                        item_elem.text = str(item)
            else:
                elem = ET.SubElement(parent, key)
                elem.text = str(value)


class StatsHTMLSerializer:
    """HTML serializer for stats responses."""

    def serialize(self, data: Union[Dict, List], **kwargs) -> Response:
        """Serialize data to HTML format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with HTML content
        """
        html = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Statistics Dashboard</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; font-weight: bold; }
                h1, h2 { color: #333; }
                .stats-container { max-width: 1200px; margin: 0 auto; }
                .stats-section { margin-bottom: 30px; }
            </style>
        </head>
        <body>
            <div class="stats-container">
                <h1>Statistics Dashboard</h1>
        """

        if isinstance(data, list) and data:
            # Handle list of records
            html += self._list_to_html(data, "Statistics Data")
        elif isinstance(data, dict):
            # Handle dict response
            html += self._dict_to_html(data)

        html += """
            </div>
        </body>
        </html>
        """

        return Response(
            html,
            mimetype="text/html",
            headers={"Content-Type": "text/html; charset=utf-8"},
        )

    def _list_to_html(self, data: List[Dict[str, Any]], title: str) -> str:
        """Convert list of records to HTML table.

        Args:
            data: List of records
            title: Table title

        Returns:
            HTML string
        """
        if not data:
            return (
                f"<div class='stats-section'><h2>{title}</h2>"
                f"<p>No data available</p></div>"
            )

        html = f"<div class='stats-section'><h2>{title}</h2><table>"

        # Header row
        fieldnames = self._get_fieldnames(data[0])
        html += "<tr>"
        for field in fieldnames:
            html += f"<th>{field}</th>"
        html += "</tr>"

        # Data rows
        for item in data:
            html += "<tr>"
            for field in fieldnames:
                value = self._get_nested_value(item, field)
                html += f"<td>{value}</td>"
            html += "</tr>"

        html += "</table></div>"
        return html

    def _dict_to_html(self, data: Dict[str, Any]) -> str:
        """Convert dictionary to HTML.

        Args:
            data: Dictionary to convert

        Returns:
            HTML string
        """
        html = ""

        for key, value in data.items():
            if isinstance(value, list) and value and isinstance(value[0], dict):
                # List of records - render as table
                html += self._list_to_html(value, key.replace("_", " ").title())
            elif isinstance(value, dict):
                # Nested dict - render as table
                html += (
                    f"<div class='stats-section'>"
                    f"<h2>{key.replace('_', ' ').title()}</h2>"
                )
                html += "<table>"
                for subkey, subvalue in value.items():
                    html += f"<tr><td>{subkey}</td><td>{subvalue}</td></tr>"
                html += "</table></div>"
            else:
                # Simple key-value pair
                html += (
                    f"<div class='stats-section'>"
                    f"<h2>{key.replace('_', ' ').title()}</h2>"
                )
                html += f"<p><strong>{key}:</strong> {value}</p></div>"

        return html

    def _get_fieldnames(self, record: Dict[str, Any]) -> List[str]:
        """Extract fieldnames from a record, handling nested structures.

        Args:
            record: Dictionary record to extract fieldnames from

        Returns:
            List of field names
        """
        fieldnames = []
        for key, value in record.items():
            if isinstance(value, dict):
                # For nested dicts, flatten with dot notation
                for subkey in value.keys():
                    fieldnames.append(f"{key}.{subkey}")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                # For lists of dicts, use the key as fieldname
                fieldnames.append(key)
            else:
                fieldnames.append(key)
        return fieldnames

    def _get_nested_value(self, record: Dict[str, Any], field: str) -> str:
        """Get value from nested field using dot notation.

        Args:
            record: Dictionary record
            field: Field name (may contain dots for nesting)

        Returns:
            String value
        """
        if "." in field:
            keys = field.split(".")
            value = record
            for key in keys:
                if isinstance(value, dict) and key in value:
                    value = value[key]
                else:
                    return ""
            return str(value) if value is not None else ""
        else:
            return str(record.get(field, ""))


class StatsExcelSerializer:
    """Excel serializer for stats responses."""

    def serialize(self, data: Union[Dict, List], **kwargs) -> Response:
        """Serialize data to Excel format.

        Args:
            data: The data to serialize (dict or list)
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with Excel content
        """
        wb = Workbook()

        if isinstance(data, list) and data:
            # Handle list of records
            ws = wb.active
            if ws is not None and hasattr(ws, "title"):
                ws.title = "Statistics Data"
            self._list_to_excel(data, ws)
        elif isinstance(data, dict):
            # Handle dict response - create multiple sheets
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    # List of records - create sheet
                    ws = wb.create_sheet(title=str(key.replace("_", " ").title()))
                    self._list_to_excel(value, ws)
                elif isinstance(value, dict):
                    # Nested dict - create sheet
                    ws = wb.create_sheet(title=str(key.replace("_", " ").title()))
                    self._dict_to_excel(value, ws)
                else:
                    # Simple value - add to summary sheet
                    if "Summary" not in [sheet.title for sheet in wb.worksheets]:
                        ws = wb.create_sheet(title="Summary")
                    else:
                        ws = wb["Summary"]
                    ws.append([key, value])

        # Style the worksheets
        for ws in wb.worksheets:
            self._style_worksheet(ws)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output.getvalue(),
            mimetype=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            headers={
                "Content-Type": (
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                "Content-Disposition": "attachment; filename=stats.xlsx",
            },
        )

    def _list_to_excel(self, data: List[Dict[str, Any]], ws) -> None:
        """Convert list of records to Excel worksheet.

        Args:
            data: List of records
            ws: Excel worksheet
        """
        if not data:
            return

        # Get fieldnames
        fieldnames = self._get_fieldnames(data[0])

        # Write headers
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)

        # Write data
        for row, record in enumerate(data, 2):
            for col, field in enumerate(fieldnames, 1):
                value = self._get_nested_value(record, field)
                ws.cell(row=row, column=col, value=value)

    def _dict_to_excel(self, data: Dict[str, Any], ws) -> None:
        """Convert dictionary to Excel worksheet.

        Args:
            data: Dictionary to convert
            ws: Excel worksheet
        """
        for row, (key, value) in enumerate(data.items(), 1):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=str(value))

    def _get_fieldnames(self, record: Dict[str, Any]) -> List[str]:
        """Extract fieldnames from a record, handling nested structures.

        Args:
            record: Dictionary record to extract fieldnames from

        Returns:
            List of field names
        """
        fieldnames = []
        for key, value in record.items():
            if isinstance(value, dict):
                # For nested dicts, flatten with dot notation
                for subkey in value.keys():
                    fieldnames.append(f"{key}.{subkey}")
            elif isinstance(value, list) and value and isinstance(value[0], dict):
                # For lists of dicts, use the key as fieldname
                fieldnames.append(key)
            else:
                fieldnames.append(key)
        return fieldnames

    def _get_nested_value(self, record: Dict[str, Any], field: str) -> str:
        """Get value from nested field using dot notation.

        Args:
            record: Dictionary record
            field: Field name (may contain dots for nesting)

        Returns:
            String value
        """
        if "." in field:
            keys = field.split(".")
            value = record
            for key in keys:
                if isinstance(value, dict) and key in value:
                    value = value[key]
                else:
                    return ""
            return str(value) if value is not None else ""
        else:
            return str(record.get(field, ""))

    def _style_worksheet(self, ws) -> None:
        """Apply styling to Excel worksheet.

        Args:
            ws: Excel worksheet to style
        """
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
