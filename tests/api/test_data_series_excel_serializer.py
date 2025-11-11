#!/usr/bin/env python3
"""Test script for the enhanced DataSeriesExcelSerializer."""

import os
import tempfile

from openpyxl import load_workbook

from invenio_stats_dashboard.resources.serializers.data_series_serializers import (
    DataSeriesExcelSerializer,
)


class TestDataSeriesExcelSerializer:
    """Test the DataSeriesExcelSerializer functionality."""

    def test_excel_workbook_creation(self, running_app):
        """Test consolidated Excel workbook creation."""
        # Sample data with query->category->metric structure
        test_data = {
            "usage-snapshot-category": {
                "access_statuses": {
                    "data_volume": [
                        {
                            "data": [["06-01", 3072.0], ["06-02", 4096.0]],
                            "year": 2025,
                            "id": "metadata-only",
                            "label": "Metadata Only",
                            "name": "",
                            "type": "line",
                            "valueType": "number",
                        }
                    ],
                    "downloads": [
                        {
                            "data": [["06-01", 3]],
                            "year": 2025,
                            "id": "metadata-only",
                            "label": "Metadata Only",
                            "name": "",
                            "type": "line",
                            "valueType": "number",
                        },
                        {
                            "data": [["06-01", 5]],
                            "year": 2025,
                            "id": "open",
                            "label": "Open Access",
                            "name": "",
                            "type": "line",
                            "valueType": "number",
                        },
                    ],
                },
                "countries": {
                    "views": [
                        {
                            "data": [["06-01", 1]],
                            "year": 2025,
                            "id": "US",
                            "label": "United States",
                            "name": "",
                            "type": "line",
                            "valueType": "number",
                        }
                    ]
                },
            }
        }

        serializer = DataSeriesExcelSerializer()

        # Test the workbook creation
        with tempfile.TemporaryDirectory() as temp_dir:
            serializer._create_excel_workbooks(test_data, temp_dir)

            # Check query folder created
            assert os.path.exists(os.path.join(temp_dir, "usage-snapshot-category"))

            # Check workbooks created (one per category within query folder)
            assert os.path.exists(
                os.path.join(
                    temp_dir, "usage-snapshot-category", "access_statuses.xlsx"
                )
            )
            assert os.path.exists(
                os.path.join(temp_dir, "usage-snapshot-category", "countries.xlsx")
            )

            # Load and verify access_statuses workbook
            wb = load_workbook(
                os.path.join(
                    temp_dir, "usage-snapshot-category", "access_statuses.xlsx"
                )
            )
            sheet_names = wb.sheetnames
            assert "data_volume" in sheet_names
            assert "downloads" in sheet_names

            # Verify data_volume sheet has correct structure
            ws = wb["data_volume"]
            assert ws.cell(row=1, column=1).value == "id"
            assert ws.cell(row=1, column=2).value == "label"
            assert ws.cell(row=1, column=3).value == "date"
            assert ws.cell(row=1, column=4).value == "value"
            assert ws.cell(row=1, column=5).value == "units"

            # Verify data rows
            assert ws.cell(row=2, column=1).value == "metadata-only"
            assert ws.cell(row=2, column=2).value == "Metadata Only"
            assert ws.cell(row=2, column=3).value == "06-01"
            assert ws.cell(row=2, column=4).value == 3072.0
            assert ws.cell(row=2, column=5).value == "bytes"

            # Verify downloads sheet consolidates multiple series
            ws = wb["downloads"]
            assert ws.cell(row=2, column=1).value == "metadata-only"
            assert ws.cell(row=2, column=5).value == "unique downloads"
            assert ws.cell(row=3, column=1).value == "open"
            assert ws.cell(row=3, column=2).value == "Open Access"

    def test_sheet_name_sanitization(self, running_app):
        """Test sheet name sanitization for Excel compatibility."""
        serializer = DataSeriesExcelSerializer()

        # Test various problematic sheet names
        test_cases = [
            ("normal_name", "normal_name"),
            (
                "name/with\\invalid?chars*[brackets]",
                "name_with_invalid_chars__bracke",
            ),
            (
                "very_long_sheet_name_that_exceeds_excel_limit_of_31_chars",
                "very_long_sheet_name_that_excee",
            ),
            ("", "Sheet"),
            ("name with spaces", "name with spaces"),
        ]

        for input_name, expected in test_cases:
            result = serializer._sanitize_sheet_name(input_name)
            assert result == expected
            assert len(result) <= 31
            invalid_chars = ["\\", "/", "?", "*", "[", "]"]
            assert not any(char in result for char in invalid_chars)

    def test_consolidated_data_addition(self):
        """Test adding consolidated data to Excel sheets."""
        serializer = DataSeriesExcelSerializer()

        test_data_series = [
            {
                "id": "en",
                "label": "English",
                "data": [["06-01", 100], ["06-02", 200]],
                "year": 2025,
            },
            {
                "id": "fr",
                "label": "French",
                "data": [["06-01", 50]],
                "year": 2025,
            },
        ]

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        assert ws is not None  # Type guard for linter
        serializer._add_consolidated_data_to_sheet(ws, "views", test_data_series)

        # Verify header
        assert ws.cell(row=1, column=1).value == "id"
        assert ws.cell(row=1, column=2).value == "label"
        assert ws.cell(row=1, column=3).value == "date"
        assert ws.cell(row=1, column=4).value == "value"
        assert ws.cell(row=1, column=5).value == "units"

        # Verify first series data
        assert ws.cell(row=2, column=1).value == "en"
        assert ws.cell(row=2, column=2).value == "English"
        assert ws.cell(row=2, column=3).value == "06-01"
        assert ws.cell(row=2, column=4).value == 100
        assert ws.cell(row=2, column=5).value == "unique views"

        # Verify second data point from first series
        assert ws.cell(row=3, column=1).value == "en"
        assert ws.cell(row=3, column=3).value == "06-02"
        assert ws.cell(row=3, column=4).value == 200

        # Verify second series data
        assert ws.cell(row=4, column=1).value == "fr"
        assert ws.cell(row=4, column=2).value == "French"
        assert ws.cell(row=4, column=3).value == "06-01"
        assert ws.cell(row=4, column=4).value == 50

    def test_empty_data_handling(self, running_app):
        """Test handling of empty or completely missing data."""
        serializer = DataSeriesExcelSerializer()

        # Test with completely empty data
        with tempfile.TemporaryDirectory() as temp_dir:
            serializer._create_excel_workbooks({}, temp_dir)
            # Should not crash and create no files
            assert len(os.listdir(temp_dir)) == 0

        # Test with data that has no actual data series
        # but valid structure
        empty_category_data: dict = {
            "usage-snapshot": {"languages": {"views": [], "downloads": []}}
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            serializer._create_excel_workbooks(empty_category_data, temp_dir)
            # Should create a workbook with "No Data" sheet
            assert os.path.exists(
                os.path.join(temp_dir, "usage-snapshot", "languages.xlsx")
            )

            wb = load_workbook(
                os.path.join(temp_dir, "usage-snapshot", "languages.xlsx")
            )
            assert "No Data" in wb.sheetnames
            ws = wb["No Data"]
            assert "No Data Available" in str(ws.cell(row=1, column=1).value)
            assert "languages" in str(ws.cell(row=2, column=1).value).lower()

    def test_partial_empty_data(self, running_app):
        """Test handling categories with both empty and populated metrics."""
        serializer = DataSeriesExcelSerializer()

        test_data = {
            "usage-snapshot": {
                "languages": {
                    "views": [
                        {
                            "id": "en",
                            "label": "English",
                            "data": [["06-01", 10]],
                            "year": 2025,
                        }
                    ],
                    "downloads": [],  # Empty metric
                }
            }
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            serializer._create_excel_workbooks(test_data, temp_dir)

            # Should create workbook with only the views sheet
            wb = load_workbook(
                os.path.join(temp_dir, "usage-snapshot", "languages.xlsx")
            )
            assert "views" in wb.sheetnames
            assert "downloads" not in wb.sheetnames
            assert "No Data" not in wb.sheetnames
