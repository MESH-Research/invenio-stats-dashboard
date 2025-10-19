# Part of the Invenio-Stats-Dashboard extension for InvenioRDM
# Copyright (C) 2025 Mesh Research
#
# Invenio-Stats-Dashboard is free software; you can redistribute it and/or modify
# it under the terms of the MIT License; see LICENSE file for more details.

# type: ignore
"""Enhanced serializers for data series with compression support."""

import csv
import gzip
import json
import os
import shutil
import tempfile
import xml.etree.ElementTree as ET

import arrow
import brotli
from flask import Response, current_app, g
from invenio_communities.proxies import current_communities
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from werkzeug.utils import secure_filename

from ...transformers.base import DataSeries


class CompressedStatsJSONSerializer:
    """Compressed JSON serializer for data series responses."""

    def __init__(self, compression_method: str | None = None):
        """Initialize the serializer with a compression method.

        Current supported compression methods are: gzip, brotli.
        The default is set in the configuration file.

        Args:
            compression_method: Compression method to use
        """
        if compression_method is None:
            compression_method = current_app.config.get(
                "STATS_CACHE_COMPRESSION_METHOD", "brotli"
            )
        self.compression_method = compression_method.lower()

    def serialize(self, data: DataSeries | dict | list, **kwargs) -> bytes:
        """Serialize data to compressed JSON format.

        Args:
            data: DataSeries object, dict, or list to serialize
            **kwargs: Additional keyword arguments

        Returns:
            Compressed JSON data as bytes
        """
        if isinstance(data, DataSeries):
            json_data = data.for_json()
        elif isinstance(data, dict):
            json_data = {}
            for key, value in data.items():
                if isinstance(value, DataSeries):
                    json_data[key] = value.for_json()
                else:
                    json_data[key] = value
        else:
            json_data = data

        json_str = json.dumps(json_data, indent=2, default=str)

        if self.compression_method == "brotli":
            compressed_data = brotli.compress(json_str.encode("utf-8"))
        else:
            compressed_data = gzip.compress(json_str.encode("utf-8"))

        return compressed_data

    def _create_gzip_response(self, compressed_data: bytes) -> Response:
        """Create a gzip-compressed response."""
        return Response(
            compressed_data,
            mimetype="application/json",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Content-Encoding": "gzip",
                "Content-Disposition": "attachment; filename=stats.json.gz",
            },
        )

    def _create_brotli_response(self, compressed_data: bytes) -> Response:
        """Create a brotli-compressed response."""
        return Response(
            compressed_data,
            mimetype="application/json",
            headers={
                "Content-Type": "application/json; charset=utf-8",
                "Content-Encoding": "br",
                "Content-Disposition": "attachment; filename=stats.json.br",
            },
        )


# Convenience classes for backward compatibility and easy configuration
class GzipStatsJSONSerializer(CompressedStatsJSONSerializer):
    """Gzip-compressed JSON serializer for data series responses."""

    def __init__(self):
        """Initialize the compressed JSON serializer with gzip compression."""
        super().__init__(compression_method="gzip")


class BrotliStatsJSONSerializer(CompressedStatsJSONSerializer):
    """Brotli-compressed JSON serializer for data series responses."""

    def __init__(self):
        """Initialize the compressed JSON serializer with brotli compression."""
        super().__init__(compression_method="brotli")


class DataSeriesCSVSerializer:
    """CSV serializer for data series responses with nested folder structure."""

    def serialize(
        self, data: DataSeries | dict | list, community_id: str | None = None, **kwargs
    ) -> bytes:
        """Serialize nested dictionary data to compressed CSV folder structure.

        Creates a temporary nested folder structure with consolidated CSV files.
        Each CSV file contains all data points for a specific metric, with columns
        for id, label, date, value, and units to distinguish between different
        subcount items and clarify the unit of measurement.

        Args:
            data: Nested dictionary with structure like sample_usage_delta_data_series
            community_id: Optional community ID for community-specific stats
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with gzip-compressed tar archive containing CSV files
        """
        if not isinstance(data, dict):
            raise ValueError("Cannot serialize non-dictionary content")

        with tempfile.TemporaryDirectory() as temp_dir:
            self._create_nested_csv_structure(data, temp_dir)

            filename_prefix = self._get_filename_prefix(community_id)
            archive_path = shutil.make_archive(
                os.path.join(tempfile.gettempdir(), filename_prefix), "gztar", temp_dir
            )

            with open(archive_path, "rb") as f:
                compressed_data = f.read()

            # Clean up the archive file
            os.unlink(archive_path)

            return compressed_data

    def _get_filename_prefix(
        self, community_id: str | None = None, format_type: str = "csv"
    ) -> str:
        """Generate filename prefix based on community and format.

        Args:
            community_id: Optional community ID
            format_type: Format type (csv, excel, xml)

        Returns:
            Filename prefix string
        """
        if community_id:
            # Get community metadata for better filename
            community_metadata = self._get_community_metadata(community_id)
            if community_metadata and community_metadata.get("slug"):
                community_slug = community_metadata["slug"]
                return f"data_series_{format_type}_{community_slug}"
            else:
                # Fallback to community ID if no metadata available
                safe_id = secure_filename(str(community_id))
                return f"data_series_{format_type}_{safe_id}"
        else:
            return f"data_series_{format_type}"

    def _get_community_metadata(self, community_id: str) -> dict | None:
        """Get community metadata from the communities service.

        Args:
            community_id: Community identifier

        Returns:
            Dictionary with community metadata or None if not found
        """
        try:
            # Get the community using the service
            community_result = current_communities.service.read(
                id_=community_id, identity=g.identity
            )
            community_data = community_result.data

            # Extract relevant metadata
            metadata = community_data.get("metadata", {})
            links = community_data.get("links", {})

            # Get the community URL
            community_url = links.get("self_html", "")
            if not community_url:
                # Construct URL from site configuration
                site_ui_url = current_app.config.get("SITE_UI_URL", "")
                slug = metadata.get("slug", community_id)
                community_url = f"{site_ui_url}/communities/{slug}"

            return {
                "title": metadata.get("title", ""),
                "description": metadata.get("description", ""),
                "slug": metadata.get("slug", ""),
                "url": community_url,
            }
        except Exception as e:
            current_app.logger.warning(
                f"Failed to retrieve community metadata for {community_id}: {e}"
            )
            return None

    def _create_nested_csv_structure(self, data: dict, base_path: str) -> None:
        """Create nested folder structure with consolidated CSV files.

        Args:
            data: Nested dictionary data with structure:
                  {series_set_name: {category_name:
                   {metric_name: [data_series_objects]}}}
            base_path: Base path for the temporary directory
        """
        # Level 0: Series sets (e.g., "record-delta-category",
        # "usage-snapshot-category")
        for series_set_name, series_set_data in data.items():
            if not isinstance(series_set_data, dict):
                continue

            # Sanitize series set name and create directory
            safe_series_set_name = secure_filename(str(series_set_name))
            series_set_path = os.path.join(base_path, safe_series_set_name)
            os.makedirs(series_set_path, exist_ok=True)

            # Level 1: Categories (e.g., "periodicals", "publishers", "affiliations")
            for category_name, category_data in series_set_data.items():
                if not isinstance(category_data, dict):
                    continue

                # Sanitize category name and create directory
                safe_category_name = secure_filename(str(category_name))
                category_path = os.path.join(series_set_path, safe_category_name)
                os.makedirs(category_path, exist_ok=True)

                # Level 2: Metrics
                # (e.g., "data_volume", "file_count", "records")
                for metric_name, metric_data in category_data.items():
                    if isinstance(metric_data, list):
                        # Create consolidated CSV for all series in metric
                        self._create_consolidated_csv_file(
                            metric_name, metric_data, category_path
                        )

    def _create_consolidated_csv_file(
        self, metric_name: str, data_series_list: list, directory_path: str
    ) -> None:
        """Create a single consolidated CSV file for all data series in a metric.

        Args:
            metric_name: Name of the metric
            data_series_list: List of data series objects
            directory_path: Directory where CSV file should be created
        """
        if not data_series_list:
            return

        safe_metric_name = secure_filename(str(metric_name))
        csv_filename = f"{safe_metric_name}.csv"
        csv_path = os.path.join(directory_path, csv_filename)

        # Get the unit for this metric
        unit = self._get_metric_unit(metric_name)
        if unit is None:
            unit = ""  # Default to empty string if no unit found

        # Collect all data points from all series
        all_rows = []

        for data_series in data_series_list:
            if not isinstance(data_series, dict):
                continue

            series_id = data_series.get("id", "unknown")
            series_label = data_series.get("label", "")

            # Handle label which might be a dict
            # (e.g., {"en": "English"})
            if isinstance(series_label, dict):
                # Try to get English label, fallback to first
                # available
                series_label = series_label.get(
                    "en", next(iter(series_label.values()), "")
                )

            data_points = data_series.get("data", [])

            if not isinstance(data_points, list):
                continue

            for data_point in data_points:
                if isinstance(data_point, dict):
                    # Extract date and value from data_point
                    value_array = data_point.get("value", [])
                    if (
                        isinstance(value_array, list)
                        and len(value_array) >= 2
                    ):
                        date_val = value_array[0]
                        numeric_val = value_array[1]
                        all_rows.append(
                            [series_id, series_label, date_val,
                             numeric_val, unit]
                        )

        # Only write file if we have data
        if not all_rows:
            return

        # Write consolidated CSV file
        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            csvwriter = csv.writer(csvfile)
            # Header row
            csvwriter.writerow(["id", "label", "date", "value", "units"])
            csvwriter.writerows(all_rows)

    def _get_metric_unit(self, metric_name: str) -> str | None:
        """Get unit for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Unit string or None
        """
        unit_mapping = {
            "data_volume": "bytes",
            "downloads": "unique downloads",
            "views": "unique views",
            "download_unique_files": "unique files downloaded",
            "download_unique_parents": "unique parents of downloaded files",
            "download_unique_records": "unique records downloaded",
            "view_unique_parents": "unique parents of viewed records",
            "view_unique_records": "unique records viewed",
            "download_visitors": "unique visitors who downloaded",
            "view_visitors": "unique visitors who viewed",
            "records": "records",
            "file_count": "files",
            "parents": "parent records",
            "uploaders": "unique uploaders",
        }
        return unit_mapping.get(metric_name.lower())


class DataSeriesExcelSerializer:
    """Enhanced Excel serializer for data series responses with multiple workbooks."""

    def serialize(
        self, data: DataSeries | dict | list, community_id: str | None = None, **kwargs
    ) -> bytes:
        """Serialize nested dictionary data to compressed Excel workbook archive.

        Creates separate Excel workbooks for each category, with one sheet per
        metric. Each sheet contains consolidated data with id, label, date, value,
        and units columns.

        Args:
            data: Nested dictionary with structure like sample_usage_delta_data_series
            community_id: Optional community ID for community-specific stats
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with gzip-compressed tar archive containing Excel files
        """
        if not isinstance(data, dict):
            raise ValueError("Cannot serialize non-dictionary content")

        # Create temporary directory for the Excel workbooks
        with tempfile.TemporaryDirectory() as temp_dir:
            self._create_excel_workbooks(data, temp_dir)

            filename_prefix = self._get_filename_prefix(community_id, "excel")
            archive_path = shutil.make_archive(
                os.path.join(tempfile.gettempdir(), filename_prefix), "gztar", temp_dir
            )

            with open(archive_path, "rb") as f:
                compressed_data = f.read()

            os.unlink(archive_path)

            return compressed_data

    def _create_excel_workbooks(self, data: dict, base_path: str) -> None:
        """Create Excel workbooks for each series set with sheets per metric.

        Args:
            data: Nested dictionary with structure:
                  {query_type: {series_set: {metric: [data_series_objects]}}}
            base_path: Base path for the temporary directory
        """
        # Create folder structure preserving query type information
        for query_name, query_data in data.items():
            if not isinstance(query_data, dict):
                continue

            # Sanitize query name and create directory
            safe_query_name = secure_filename(str(query_name))
            query_path = os.path.join(base_path, safe_query_name)
            os.makedirs(query_path, exist_ok=True)

            # Create one workbook per series set within the query folder
            for series_set_name, series_set_data in query_data.items():
                if not isinstance(series_set_data, dict):
                    continue

                # Create workbook for this series set
                self._create_series_set_workbook(series_set_name, series_set_data, query_path)

    def _create_series_set_workbook(self, series_set_name: str, series_set_data: dict, query_path: str) -> None:
        """Create a single Excel workbook for a series set with sheets per metric.

        Args:
            series_set_name: Name of the series set
            series_set_data: Dictionary containing metrics data
            query_path: Path to the query folder
        """
        wb = Workbook()
        # Keep track of whether we created any sheets
        sheets_created = 0

        # Create one sheet per metric
        for metric_name, data_series_list in series_set_data.items():
            if not isinstance(data_series_list, list) or not data_series_list:
                continue

            # Remove default sheet only before creating first sheet
            if sheets_created == 0 and wb.active:
                wb.remove(wb.active)

            sheet_name = self._sanitize_sheet_name(str(metric_name))
            ws = wb.create_sheet(title=sheet_name)

            # Add consolidated data to sheet
            self._add_consolidated_data_to_sheet(
                ws, metric_name, data_series_list
            )
            sheets_created += 1

        # If no sheets were created, add a "No Data" sheet
        if sheets_created == 0:
            if wb.active:
                ws = wb.active
                ws.title = "No Data"
            else:
                ws = wb.create_sheet(title="No Data")

            # Add message
            ws.cell(row=1, column=1, value="No Data Available")
            ws.cell(
                row=2,
                column=1,
                value=f"No data available for series set: {series_set_name}",
            )
            ws.cell(
                row=3,
                column=1,
                value="This may indicate no activity during the "
                "requested period.",
            )

            # Style the message
            header_font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).font = header_font
        else:
            # Style the workbook with data
            self._style_workbook(wb)

        # Save workbook to file
        safe_series_set_name = secure_filename(str(series_set_name))
        excel_filename = f"{safe_series_set_name}.xlsx"
        excel_path = os.path.join(query_path, excel_filename)
        wb.save(excel_path)

    def _add_consolidated_data_to_sheet(
        self, ws, metric_name: str, data_series_list: list
    ) -> None:
        """Add consolidated data to Excel sheet.

        Args:
            ws: Excel worksheet
            metric_name: Name of the metric
            data_series_list: List of data series objects
        """
        # Get the unit for this metric
        unit = self._get_metric_unit(metric_name)
        if unit is None:
            unit = ""

        # Add header row
        ws.cell(row=1, column=1, value="id")
        ws.cell(row=1, column=2, value="label")
        ws.cell(row=1, column=3, value="date")
        ws.cell(row=1, column=4, value="value")
        ws.cell(row=1, column=5, value="units")

        # Collect all data points from all series
        current_row = 2

        for data_series in data_series_list:
            if not isinstance(data_series, dict):
                continue

            series_id = data_series.get("id", "unknown")
            series_label = data_series.get("label", "")

            # Handle label which might be a dict
            # (e.g., {"en": "English"})
            if isinstance(series_label, dict):
                # Try to get English label, fallback to first
                # available
                series_label = series_label.get(
                    "en", next(iter(series_label.values()), "")
                )

            data_points = data_series.get("data", [])

            if not isinstance(data_points, list):
                continue

            for data_point in data_points:
                if isinstance(data_point, dict):
                    # Extract date and value from data_point
                    value_array = data_point.get("value", [])
                    if (
                        isinstance(value_array, list)
                        and len(value_array) >= 2
                    ):
                        date_val = value_array[0]
                        numeric_val = value_array[1]

                        # Add row with all columns
                        ws.cell(row=current_row, column=1, value=series_id)
                        ws.cell(
                            row=current_row, column=2, value=series_label
                        )
                        ws.cell(row=current_row, column=3, value=date_val)
                        ws.cell(
                            row=current_row, column=4, value=numeric_val
                        )
                        ws.cell(row=current_row, column=5, value=unit)
                        current_row += 1

    def _sanitize_sheet_name(self, name: str) -> str:
        """Sanitize sheet name for Excel compatibility.

        Args:
            name: Original sheet name

        Returns:
            Sanitized sheet name
        """
        # Excel sheet names have restrictions:
        # - Max 31 characters
        # - Cannot contain: \ / ? * [ ]
        # - Cannot be empty
        sanitized = str(name)

        # Remove invalid characters
        invalid_chars = ["\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")

        # Truncate if too long
        if len(sanitized) > 31:
            sanitized = sanitized[:31]

        # Ensure not empty
        if not sanitized:
            sanitized = "Sheet"

        return sanitized

    def _style_workbook(self, wb: Workbook) -> None:
        """Apply styling to workbook.

        Args:
            wb: Excel workbook
        """
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        for ws in wb.worksheets:
            # Style header row
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

    def _get_filename_prefix(
        self, community_id: str | None = None, format_type: str = "csv"
    ) -> str:
        """Generate filename prefix based on community and format.

        Args:
            community_id: Optional community ID
            format_type: Format type (csv, excel, xml)

        Returns:
            Filename prefix string
        """
        if community_id:
            # Get community metadata for better filename
            community_metadata = self._get_community_metadata(community_id)
            if community_metadata and community_metadata.get("slug"):
                community_slug = community_metadata["slug"]
                return f"data_series_{format_type}_{community_slug}"
            else:
                # Fallback to community ID if no metadata available
                safe_id = secure_filename(str(community_id))
                return f"data_series_{format_type}_{safe_id}"
        else:
            return f"data_series_{format_type}"

    def _get_community_metadata(self, community_id: str) -> dict | None:
        """Get community metadata from the communities service.

        Args:
            community_id: Community identifier

        Returns:
            Dictionary with community metadata or None if not found
        """
        try:
            # Get the community using the service
            community_result = current_communities.service.read(
                id_=community_id, identity=g.identity
            )
            community_data = community_result.data

            # Extract relevant metadata
            metadata = community_data.get("metadata", {})
            links = community_data.get("links", {})

            # Get the community URL
            community_url = links.get("self_html", "")
            if not community_url:
                # Construct URL from site configuration
                site_ui_url = current_app.config.get("SITE_UI_URL", "")
                slug = metadata.get("slug", community_id)
                community_url = f"{site_ui_url}/communities/{slug}"

            return {
                "title": metadata.get("title", ""),
                "description": metadata.get("description", ""),
                "slug": metadata.get("slug", ""),
                "url": community_url,
            }
        except Exception as e:
            current_app.logger.warning(
                f"Failed to retrieve community metadata for "
                f"{community_id}: {e}"
            )
            return None

    def _get_metric_unit(self, metric_name: str) -> str | None:
        """Get unit for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Unit string or None
        """
        unit_mapping = {
            "data_volume": "bytes",
            "downloads": "unique downloads",
            "views": "unique views",
            "download_unique_files": "unique files downloaded",
            "download_unique_parents": "unique parents of downloaded files",
            "download_unique_records": "unique records downloaded",
            "view_unique_parents": "unique parents of viewed records",
            "view_unique_records": "unique records viewed",
            "download_visitors": "unique visitors who downloaded",
            "view_visitors": "unique visitors who viewed",
            "records": "records",
            "file_count": "files",
            "parents": "parent records",
            "uploaders": "unique uploaders",
        }
        return unit_mapping.get(metric_name.lower())


class DataSeriesXMLSerializer:
    """XML serializer for data series responses."""

    def serialize(
        self, data: DataSeries | dict | list, community_id: str | None = None, **kwargs
    ) -> str:
        """Serialize nested dictionary data to structured XML format.

        Creates a well-structured XML representation of the nested data with
        proper elements for series categories, metrics, and data points.

        Args:
            data: Nested dictionary with structure like sample_usage_delta_data_series
            community_id: Optional community ID for community-specific stats
            **kwargs: Additional keyword arguments

        Returns:
            Flask Response with XML content
        """
        if not isinstance(data, dict):
            raise ValueError("Cannot serialize non-dictionary content")

        # Create root element with namespace and schema reference
        root = ET.Element("dataSeriesCollection")
        root.set("xmlns", "https://github.com/MESH-Research/invenio-stats-dashboard")
        root.set("xmlns:dc", "http://purl.org/dc/elements/1.1/")
        root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        root.set(
            "xsi:schemaLocation",
            "https://github.com/MESH-Research/invenio-stats-dashboard "
            "https://github.com/MESH-Research/invenio-stats-dashboard/"
            "schema/data-series.xsd",
        )
        root.set("version", "1.0")

        # Add comprehensive metadata with Dublin Core elements
        metadata = ET.SubElement(root, "metadata")

        publisher = self._get_publisher_from_config()
        # Dublin Core metadata
        ET.SubElement(metadata, "dc:title").text = f"{publisher} Data Series Collection"
        ET.SubElement(metadata, "dc:creator").text = publisher
        ET.SubElement(metadata, "dc:description").text = (
            "Time-series statistical data from the "
            f"{publisher} "
            "including usage metrics and record counts."
        )
        ET.SubElement(metadata, "dc:date").text = arrow.utcnow().isoformat()
        ET.SubElement(metadata, "dc:format").text = "application/xml"
        # Get language from i18n configuration
        language = self._get_language_from_config()
        ET.SubElement(metadata, "dc:language").text = language
        ET.SubElement(metadata, "dc:publisher").text = publisher
        ET.SubElement(metadata, "dc:source").text = f"{publisher} API"
        ET.SubElement(
            metadata, "dc:subject"
        ).text = "Statistics, Usage Analytics, Time Series Data"
        ET.SubElement(metadata, "dc:type").text = "Dataset"

        # Technical metadata
        ET.SubElement(metadata, "generatedAt").text = arrow.utcnow().isoformat()
        ET.SubElement(metadata, "totalCategories").text = str(len(data))

        # Calculate total data points
        total_points = self._calculate_total_data_points(data)
        ET.SubElement(metadata, "totalDataPoints").text = str(total_points)

        # Add time range if available
        time_range = self._calculate_time_range(data)
        if time_range:
            time_range_elem = ET.SubElement(metadata, "timeRange")
            ET.SubElement(time_range_elem, "startDate").text = time_range[0]
            ET.SubElement(time_range_elem, "endDate").text = time_range[1]

        # Add community metadata if community_id is provided
        if community_id:
            community_metadata = self._get_community_metadata(community_id)
            if community_metadata:
                community_elem = ET.SubElement(metadata, "community")
                ET.SubElement(community_elem, "dc:identifier").text = community_id
                ET.SubElement(community_elem, "dc:title").text = community_metadata.get(
                    "title", ""
                )
                ET.SubElement(
                    community_elem, "dc:description"
                ).text = community_metadata.get("description", "")
                ET.SubElement(
                    community_elem, "dc:source"
                ).text = community_metadata.get("url", "")
                ET.SubElement(community_elem, "slug").text = community_metadata.get(
                    "slug", ""
                )

        # Process each top-level query type (category)
        for query_name, query_data in data.items():
            category_elem = ET.SubElement(root, "category")
            category_elem.set("name", str(query_name))
            category_elem.set("id", self._sanitize_xml_id(str(query_name)))

            # Add semantic attributes for category
            category_type = self._get_category_type(str(query_name))
            if category_type:
                category_elem.set("categoryType", category_type)

            description = self._get_category_description(str(query_name))
            if description:
                category_elem.set("description", description)

            # Count total metrics across all series sets
            total_metrics_count = 0
            if isinstance(query_data, dict):
                for series_set_data in query_data.values():
                    if isinstance(series_set_data, dict):
                        total_metrics_count += sum(
                            1 for v in series_set_data.values() if isinstance(v, list) and v
                        )
            category_elem.set("metricsCount", str(total_metrics_count))

            # Process each series set within the category (only if data is valid)
            if isinstance(query_data, dict):
                for series_set_name, series_set_data in query_data.items():
                    if not isinstance(series_set_data, dict):
                        continue

                    series_set_elem = ET.SubElement(category_elem, "seriesSet")
                    series_set_elem.set("name", str(series_set_name))
                    series_set_elem.set("id", self._sanitize_xml_id(str(series_set_name)))

                    # Count metrics in this series set
                    series_set_metrics_count = sum(
                        1 for v in series_set_data.values() if isinstance(v, list) and v
                    )
                    series_set_elem.set("metricsCount", str(series_set_metrics_count))

                    # Process each metric in the series set
                    for metric_name, metric_data in series_set_data.items():
                        if not isinstance(metric_data, list):
                            continue

                        # Count total data points across all series in this metric
                        total_data_points = 0
                        for series_obj in metric_data:
                            if isinstance(series_obj, dict) and "data" in series_obj:
                                series_data = series_obj["data"]
                                if isinstance(series_data, list):
                                    total_data_points += len(series_data)

                        metric_elem = ET.SubElement(series_set_elem, "metric")
                        metric_elem.set("name", str(metric_name))
                        metric_elem.set("id", self._sanitize_xml_id(str(metric_name)))
                        metric_elem.set("dataPointsCount", str(total_data_points))

                        # Add semantic attributes for metric
                        unit = self._get_metric_unit(str(metric_name))
                        if unit:
                            metric_elem.set("unit", unit)

                        measurement_type = self._get_measurement_type(str(metric_name))
                        if measurement_type:
                            metric_elem.set("measurementType", measurement_type)

                        aggregation_method = self._get_aggregation_method(str(metric_name))
                        if aggregation_method:
                            metric_elem.set("aggregationMethod", aggregation_method)

                        description = self._get_metric_description(str(metric_name))
                        if description:
                            metric_elem.set("description", description)

                        # Process each data series in the metric
                        for series_obj in metric_data:
                            if not isinstance(series_obj, dict) or "id" not in series_obj:
                                continue

                            series_elem = ET.SubElement(metric_elem, "series")
                            series_elem.set("id", str(series_obj.get("id", "unknown")))

                            # Add series metadata
                            if "name" in series_obj:
                                series_elem.set("name", str(series_obj["name"]))
                            if "type" in series_obj:
                                series_elem.set("type", str(series_obj["type"]))
                            if "valueType" in series_obj:
                                series_elem.set("valueType", str(series_obj["valueType"]))

                            # Add semantic attributes
                            if "label" in series_obj:
                                series_elem.set("label", str(series_obj["label"]))

                            description = self._get_series_description(series_obj)
                            if description:
                                series_elem.set("description", description)

                            # Add data points
                            data_points = series_obj.get("data", [])
                            if isinstance(data_points, list):
                                points_elem = ET.SubElement(series_elem, "dataPoints")
                                points_elem.set("count", str(len(data_points)))

                                for point in data_points:
                                    if isinstance(point, dict):
                                        point_elem = ET.SubElement(points_elem, "point")

                                        # Add readable date if available
                                        if "readableDate" in point:
                                            point_elem.set(
                                                "readableDate", str(point["readableDate"])
                                            )

                                        # Add value array
                                        value_array = point.get("value", [])
                                        if (
                                            isinstance(value_array, list)
                                            and len(value_array) >= 2
                                        ):
                                            point_elem.set("date", str(value_array[0]))
                                            point_elem.set("value", str(value_array[1]))

                                            # Add value type if available
                                            if "valueType" in point:
                                                point_elem.set(
                                                    "valueType", str(point["valueType"])
                                                )

                                            # Add semantic attributes
                                            unit = self._get_metric_unit(str(metric_name))
                                            if unit:
                                                point_elem.set("unit", unit)

                                            # Add quality indicator
                                            quality = self._assess_data_quality(point)
                                            if quality:
                                                point_elem.set("quality", quality)

        # Format XML with proper indentation
        self._indent_xml(root)

        # Create XML string
        xml_string = ET.tostring(root, encoding="unicode", xml_declaration=True)

        # Generate community-specific filename
        # filename_prefix = self._get_filename_prefix(community_id, "xml")
        return xml_string

    def _get_filename_prefix(
        self, community_id: str | None = None, format_type: str = "xml"
    ) -> str:
        """Generate filename prefix based on community and format.

        Args:
            community_id: Optional community ID
            format_type: Format type (csv, excel, xml)

        Returns:
            Filename prefix string
        """
        if community_id:
            # Get community metadata for better filename
            community_metadata = self._get_community_metadata(community_id)
            if community_metadata and community_metadata.get("slug"):
                community_slug = community_metadata["slug"]
                return f"data_series_{format_type}_{community_slug}"
            else:
                # Fallback to community ID if no metadata available
                safe_id = secure_filename(str(community_id))
                return f"data_series_{format_type}_{safe_id}"
        else:
            return f"data_series_{format_type}"

    def _sanitize_xml_id(self, name: str) -> str:
        """Sanitize name for use as XML ID attribute.

        Args:
            name: Original name

        Returns:
            Sanitized ID suitable for XML
        """
        # XML IDs must start with letter or underscore, contain only letters,
        # digits, hyphens, underscores, and periods
        sanitized = str(name)

        # Replace invalid characters with underscores
        invalid_chars = [
            " ",
            "/",
            "\\",
            "?",
            "*",
            "[",
            "]",
            "(",
            ")",
            "{",
            "}",
            ":",
            ";",
            "=",
            "+",
            "&",
            "%",
            "$",
            "#",
            "@",
            "!",
            "~",
            "`",
            "^",
            "|",
            "<",
            ">",
            ",",
            ".",
            '"',
            "'",
        ]
        for char in invalid_chars:
            sanitized = sanitized.replace(char, "_")

        if sanitized and not (sanitized[0].isalpha() or sanitized[0] == "_"):
            sanitized = f"id_{sanitized}"

        if not sanitized:
            sanitized = "unknown"

        return sanitized

    def _get_publisher_from_config(self) -> str:
        """Get publisher information from InvenioRDM configuration.

        Returns:
            Publisher name from configuration or default fallback
        """
        # Try to get publisher from various configuration sources
        publisher = (
            current_app.config.get("THEME_SITENAME")
            or current_app.config.get("THEME_FRONTPAGE_TITLE")
            or current_app.config.get("THEME_SHORT_TITLE")
            or current_app.config.get("APP_THEME")
            or "InvenioRDM"
        )
        return publisher

    def _get_language_from_config(self) -> str:
        """Get language from InvenioRDM i18n configuration.

        Returns:
            Language code from configuration or default fallback
        """
        # Get language from BABEL_DEFAULT_LOCALE configuration
        language = (
            current_app.config.get("BABEL_DEFAULT_LOCALE")
            or current_app.config.get("I18N_DEFAULT_LOCALE")
            or "en"
        )
        return language

    def _get_community_metadata(self, community_id: str) -> dict | None:
        """Get community metadata from the communities service.

        Args:
            community_id: Community identifier

        Returns:
            Dictionary with community metadata or None if not found
        """
        try:
            community_result = current_communities.service.read(
                id_=community_id, identity=g.identity
            )
            community_data = community_result.data

            # Extract relevant metadata
            metadata = community_data.get("metadata", {})
            links = community_data.get("links", {})

            # Get the community URL
            community_url = links.get("self_html", "")
            if not community_url:
                # Construct URL from site configuration
                site_ui_url = current_app.config.get("SITE_UI_URL", "")
                slug = metadata.get("slug", community_id)
                community_url = f"{site_ui_url}/communities/{slug}"

            return {
                "title": metadata.get("title", ""),
                "description": metadata.get("description", ""),
                "slug": metadata.get("slug", ""),
                "url": community_url,
            }
        except Exception as e:
            current_app.logger.warning(
                f"Failed to retrieve community metadata for {community_id}: {e}"
            )
            return None

    def _calculate_total_data_points(self, data: dict) -> int:
        """Calculate total number of data points across all series.

        Args:
            data: Nested dictionary data

        Returns:
            Total number of data points
        """
        total = 0
        for category_data in data.values():
            if isinstance(category_data, dict):
                for metric_data in category_data.values():
                    if isinstance(metric_data, list):
                        for series_obj in metric_data:
                            if isinstance(series_obj, dict):
                                data_points = series_obj.get("data", [])
                                if isinstance(data_points, list):
                                    total += len(data_points)
        return total

    def _calculate_time_range(self, data: dict) -> tuple[str, str] | None:
        """Calculate the time range covered by the data.

        Args:
            data: Nested dictionary data

        Returns:
            Tuple of (start_date, end_date) in ISO format, or None if no data
        """
        dates = []
        for category_data in data.values():
            if isinstance(category_data, dict):
                for metric_data in category_data.values():
                    if isinstance(metric_data, list):
                        for series_obj in metric_data:
                            if isinstance(series_obj, dict):
                                data_points = series_obj.get("data", [])
                                if isinstance(data_points, list):
                                    for point in data_points:
                                        if isinstance(point, dict):
                                            value_array = point.get("value", [])
                                            if (
                                                isinstance(value_array, list)
                                                and len(value_array) >= 2
                                            ):
                                                dates.append(value_array[0])

        if not dates:
            return None

        dates.sort()
        return (dates[0], dates[-1])

    def _get_category_type(self, category_name: str) -> str | None:
        """Get semantic category type based on category name.

        Args:
            category_name: Name of the category

        Returns:
            Category type or None
        """
        category_mapping = {
            "access_statuses": "access_status",
            "countries": "geographic",
            "languages": "content_type",
            "resource_types": "content_type",
            "subjects": "content_type",
            "publishers": "content_type",
            "rights": "content_type",
            "funders": "content_type",
            "affiliations": "geographic",
            "file_types": "content_type",
            "referrers": "user_behavior",
            "periodicals": "content_type",
            "global": "global",
        }
        return category_mapping.get(category_name.lower())

    def _get_category_description(self, category_name: str) -> str | None:
        """Get description for category based on name.

        Args:
            category_name: Name of the category

        Returns:
            Description or None
        """
        descriptions = {
            "access_statuses": (
                "Statistics grouped by access status (open, restricted, etc.)"
            ),
            "countries": "Statistics grouped by geographic country",
            "languages": "Statistics grouped by content language",
            "resource_types": "Statistics grouped by type of resource",
            "subjects": "Statistics grouped by subject classification",
            "publishers": "Statistics grouped by publisher",
            "rights": "Statistics grouped by license/rights",
            "funders": "Statistics grouped by funding organization",
            "affiliations": "Statistics grouped by institutional affiliation",
            "file_types": "Statistics grouped by file format type",
            "referrers": "Statistics grouped by referring website",
            "periodicals": "Statistics grouped by journal/publication",
            "global": "Global statistics across all categories",
        }
        return descriptions.get(category_name.lower())

    def _get_metric_unit(self, metric_name: str) -> str | None:
        """Get unit for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Unit string or None
        """
        unit_mapping = {
            "data_volume": "bytes",
            "downloads": "count",
            "views": "count",
            "download_unique_files": "count",
            "download_unique_parents": "count",
            "download_unique_records": "count",
            "view_unique_parents": "count",
            "view_unique_records": "count",
            "download_visitors": "count",
            "view_visitors": "count",
        }
        return unit_mapping.get(metric_name.lower())

    def _get_measurement_type(self, metric_name: str) -> str | None:
        """Get measurement type for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Measurement type or None
        """
        type_mapping = {
            "data_volume": "volume",
            "downloads": "count",
            "views": "count",
            "download_unique_files": "count",
            "download_unique_parents": "count",
            "download_unique_records": "count",
            "view_unique_parents": "count",
            "view_unique_records": "count",
            "download_visitors": "count",
            "view_visitors": "count",
        }
        return type_mapping.get(metric_name.lower())

    def _get_aggregation_method(self, metric_name: str) -> str | None:
        """Get aggregation method for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Aggregation method or None
        """
        # Most metrics are daily aggregations
        if "unique" in metric_name.lower():
            return "daily"
        elif "visitors" in metric_name.lower():
            return "daily"
        else:
            return "daily"

    def _get_metric_description(self, metric_name: str) -> str | None:
        """Get description for metric based on name.

        Args:
            metric_name: Name of the metric

        Returns:
            Description or None
        """
        descriptions = {
            "data_volume": "Total data volume transferred",
            "downloads": "Total number of download events",
            "views": "Total number of view events",
            "download_unique_files": "Number of unique files downloaded",
            "download_unique_parents": "Number of unique parent records downloaded",
            "download_unique_records": "Number of unique records downloaded",
            "view_unique_parents": "Number of unique parent records viewed",
            "view_unique_records": "Number of unique records viewed",
            "download_visitors": "Number of unique visitors who downloaded",
            "view_visitors": "Number of unique visitors who viewed",
        }
        return descriptions.get(metric_name.lower())

    def _get_series_description(self, series_obj: dict) -> str | None:
        """Get description for series based on object data.

        Args:
            series_obj: Series object

        Returns:
            Description or None
        """
        series_id = series_obj.get("id", "")
        if series_id == "global":
            return "Global statistics across all data"
        elif series_id == "metadata-only":
            return "Statistics for metadata-only records"
        elif series_id == "US":
            return "Statistics for United States"
        elif series_id == "eng":
            return "Statistics for English language content"
        elif series_id == "pdf":
            return "Statistics for PDF files"
        elif series_id == "google.com":
            return "Statistics for traffic from Google"
        else:
            return f"Statistics for {series_id}"

    def _assess_data_quality(self, point: dict) -> str | None:
        """Assess data quality for a data point.

        Args:
            point: Data point dictionary

        Returns:
            Quality assessment or None
        """
        # Simple quality assessment based on available data
        if "readableDate" in point and "value" in point:
            return "high"
        elif "value" in point:
            return "medium"
        else:
            return "low"

    def _indent_xml(self, elem: ET.Element, level: int = 0) -> None:
        """Add proper indentation to XML elements.

        Args:
            elem: XML element to indent
            level: Current indentation level
        """
        indent = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = indent + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = indent
            for child in elem:
                self._indent_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = indent
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = indent
